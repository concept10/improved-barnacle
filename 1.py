# Create a dictionary to store tagname, ID, and RGBA color for each tagname



import os
import hashlib
import openpyxl

# Load the Tracker spreadsheet
wb_tracker = openpyxl.load_workbook('Tracker.xlsx')
sheet_tracker = wb_tracker.active

# Create a dictionary to store tagname, ID, and RGBA color for each tagname
tag_data = {}
for row in sheet_tracker.iter_rows(values_only=True):
    tagname = row[0]
    if tagname:
        # Generate a unique ID for each tagname
        tag_id = hashlib.sha1(tagname.encode()).hexdigest()
        # Generate a unique RGBA color for each tagname
        tag_color = (hash(tagname) % 255, hash(tag_id) % 255, (hash(tagname) + hash(tag_id)) % 255, 1)
        tag_data[tagname] = (tag_id, tag_color)

# Load the PMD spreadsheet
wb_pmd = openpyxl.load_workbook('pmd.xlsx')
sheet_pmd = wb_pmd.active

# Associate tagnames in the PMD spreadsheet with those in the Tracker spreadsheet
pmd_data = {}
for row in sheet_pmd.iter_rows(values_only=True):
    tagname = row[0]
    if tagname in tag_data:
        pmd_data[tagname] = tag_data[tagname]

# Search the workDirectory for each PDF file associated with each tagname
work_directory = '/path/to/workDirectory'
for tagname, (tag_id, tag_color) in pmd_data.items():
    for filename in os.listdir(work_directory):
        if filename.startswith(tagname) and filename.endswith('.pdf'):
            # Use OCR to find the location in the PDF file
            # Generate XML output of location, color, and ID
            location = # Code to find location using OCR
            xml = f'<Tag id="{tag_id}" color="{tag_color}" location="{location}"/>'
            # Write the XML output to a file
            with open(f'{tagname}.xml', 'w') as f:
                f.write(xml)
